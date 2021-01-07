def set_hyperlink():
    from openpyxl import load_workbook
    x = 'C:/Users/IWA-LSM/Desktop/EBW/DBProgram_pyqt/input.png'
    wb = load_workbook("hello.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")
    ws.cell(row=1, column =1).hyperlink = x
    wb.save("hello.xlsx")


set_hyperlink()