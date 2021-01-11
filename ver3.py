import sys

import openpyxl
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QApplication
from PyQt5.QtWidgets import QComboBox
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QFormLayout
from PyQt5.QtWidgets import QGridLayout
from PyQt5.QtWidgets import QGroupBox
from PyQt5.QtWidgets import QHBoxLayout
from PyQt5.QtWidgets import QLabel
from PyQt5.QtWidgets import QLineEdit
from PyQt5.QtWidgets import QMainWindow
from PyQt5.QtWidgets import QPushButton
from PyQt5.QtWidgets import QTabWidget
from PyQt5.QtWidgets import QVBoxLayout
from PyQt5.QtWidgets import QWidget


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        tabs = QTabWidget()
        tabs.addTab(TabData(), 'Data')
        tabs.addTab(TabCompare(), 'Comparison')

        self.setCentralWidget(tabs)

        self.setWindowTitle('EBW DB Program by IWA')
        self.show()


class TabData(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        tapdatagrid = QGridLayout()
        tapdatagrid.addWidget(self.groupsetup(), 0, 0)
        tapdatagrid.addWidget(self.grouptestinput(), 1, 0)
        tapdatagrid.addWidget(self.groupcosmetic(), 0, 1)
        tapdatagrid.addWidget(self.grouptestoutput(), 1, 1)
        tapdatagrid.addWidget(self.groupimage(), 0, 2, 2, 2)
        tapdatagrid.addWidget(self.saveButton(), 2, 0, 1, 4)

        self.setLayout(tapdatagrid)

    def groupsetup(self):
        group_setup = QGroupBox('Setup')
        flay_setup = QFormLayout()
        group_setup.setLayout(flay_setup)

        array_use = ["Plate", "Tube"]
        self.cb_use = QComboBox()
        self.cb_use.addItems(array_use)
        self.setup_use = self.cb_use.currentText()

        arrya_weldtype = ["BOP", "Butt", "Lap"]
        self.cb_weldtype = QComboBox()
        self.cb_weldtype.addItems(arrya_weldtype)
        self.setup_weldtype = self.cb_weldtype.currentText()

        # 값을 반환해야 하는 위젯들은 미리 선언해준다.
        self.setup_name = QLineEdit()
        self.setup_company = QLineEdit()
        self.setup_material = QLineEdit()
        self.setup_diameter = QLineEdit()
        self.setup_thickness = QLineEdit()
        self.setup_backth = QLineEdit()

        flay_setup.addRow("Name: ", self.setup_name)
        flay_setup.addRow("Company: ", self.setup_company)
        flay_setup.addRow("Use: ", self.cb_use)
        flay_setup.addRow("Material: ", self.setup_material)
        flay_setup.addRow("Weld Type: ", self.cb_weldtype)
        flay_setup.addRow("Diameter(mm): ", self.setup_diameter)
        flay_setup.addRow("Thickness(mm): ", self.setup_thickness)
        flay_setup.addRow("Backing Thickness(mm): ", self.setup_backth)

        return group_setup

    def grouptestinput(self):
        group_testinput = QGroupBox('Test Input')  # groupbox 변수 정의
        flay_testinput = QFormLayout()  # groupbox 형식 변수 정의
        group_testinput.setLayout(flay_testinput)  # groupbox 변수의 형식 입력

        self.array_sc = ["o", "|", "-", "∩", "∪", "oval"]
        self.cb_inputsc = QComboBox()
        self.cb_inputsc.addItems(self.array_sc)
        self.input_sctype = self.cb_inputsc.currentText()

        hlay_sc = QHBoxLayout()
        self.input_sclength1 = QLineEdit()
        self.input_sclength2 = QLineEdit()

        hlay_sc.addWidget(self.input_sclength1)
        hlay_sc.addWidget(self.input_sclength2)

        self.input_fobs = QLineEdit()
        self.input_wd = QLineEdit()
        self.input_bw = QLineEdit()
        self.input_fw = QLineEdit()
        self.input_velocity = QLineEdit()

        flay_testinput.addRow("Fobs(mA): ", self.input_fobs)
        flay_testinput.addRow("WD(mm): ", self.input_wd)
        flay_testinput.addRow("Bw(mA): ", self.input_bw)
        flay_testinput.addRow("Fw(mA): ", self.input_fw)
        flay_testinput.addRow("Velocity(mm/min): ", self.input_velocity)
        flay_testinput.addRow("S.C(mA): ", self.cb_inputsc)
        flay_testinput.addRow(hlay_sc)
        return group_testinput

    def groupcosmetic(self):
        self.group_cosmetic = QGroupBox('Cosmetic Input')  # groupbox 변수 정의
        self.group_cosmetic.setCheckable(True)
        self.group_cosmetic.setChecked(True)

        flay_cosmetic = QFormLayout()  # groupbox 형식 변수 정의
        self.group_cosmetic.setLayout(flay_cosmetic)  # groupbox 변수의 형식 입력

        self.cb_cosmeticsc = QComboBox()
        self.cb_cosmeticsc.addItems(self.array_sc)
        self.cosmetic_sctype = self.cb_cosmeticsc.currentText()

        hlay_sc = QHBoxLayout()
        self.cosmetic_sclength1 = QLineEdit()
        self.cosmetic_sclength2 = QLineEdit()

        hlay_sc.addWidget(self.cosmetic_sclength1)
        hlay_sc.addWidget(self.cosmetic_sclength2)

        self.cosmetic_fobs = QLineEdit()
        self.cosmetic_wd = QLineEdit()
        self.cosmetic_bw = QLineEdit()
        self.cosmetic_fw = QLineEdit()
        self.cosmetic_velocity = QLineEdit()

        flay_cosmetic.addRow("Fobs(mA): ", self.cosmetic_fobs)
        flay_cosmetic.addRow("WD(mm): ", self.cosmetic_wd)
        flay_cosmetic.addRow("Bw(mA): ", self.cosmetic_bw)
        flay_cosmetic.addRow("Fw(mA): ", self.cosmetic_fw)
        flay_cosmetic.addRow("Velocity(mm/min): ", self.cosmetic_velocity)
        flay_cosmetic.addRow("S.C(mA): ", self.cb_cosmeticsc)
        flay_cosmetic.addRow(hlay_sc)
        return self.group_cosmetic

    def grouptestoutput(self):
        group_testoutput = QGroupBox("Output")
        flay_testoutput = QFormLayout()
        group_testoutput.setLayout(flay_testoutput)

        self.output_beadW = QLineEdit()
        self.output_beadH = QLineEdit()
        self.output_backW = QLineEdit()
        self.output_backH = QLineEdit()
        self.output_50W = QLineEdit()
        self.output_50H = QLineEdit()
        self.output_crack = QLineEdit()
        self.output_porosity = QLineEdit()

        flay_testoutput.addRow("Bead Width(mm): ", self.output_beadW)
        flay_testoutput.addRow("Bead Height(mm): ", self.output_beadH)
        flay_testoutput.addRow("Back Width(mm): ", self.output_backW)
        flay_testoutput.addRow("Back Height(mm): ", self.output_backH)
        flay_testoutput.addRow("50% Width(mm): ", self.output_50W)
        flay_testoutput.addRow("50% Height(mm): ", self.output_50H)
        flay_testoutput.addRow("Crack(mm): ", self.output_crack)
        flay_testoutput.addRow("Porosity(mm): ", self.output_porosity)

        return group_testoutput

    def groupimage(self):
        group_image = QGroupBox("Image")
        flay_image = QFormLayout()
        group_image.setLayout(flay_image)

        self.btn_openimage = QPushButton("Open")
        self.lb_beadimage = QLabel()

        flay_image.addWidget(self.btn_openimage)
        flay_image.addWidget(self.lb_beadimage)
        self.lb_beadimage.setFixedSize(500, 500)

        self.btn_openimage.clicked.connect(self.openimageClicked)

        return group_image

    def openimageClicked(self):
        fname_openimage = QFileDialog.getOpenFileName(self, 'Open File')
        self.fname_openimage = fname_openimage[0]
        pixmap_bead = QPixmap(self.fname_openimage)

        self.lb_beadimage.setPixmap(pixmap_bead.scaled(self.lb_beadimage.width(), self.lb_beadimage.height()))

    def saveButton(self):
        group_save = QGroupBox("Save")

        btnEx = QPushButton('Existing File')
        btnNew = QPushButton('New File')

        btnEx.clicked.connect(self.SaveExsitantFile)
        btnNew.clicked.connect(self.SaveNewFile)

        hbox_save = QHBoxLayout()
        hbox_save.addWidget(btnEx)
        hbox_save.addWidget(btnNew)

        group_save.setLayout(hbox_save)

        return group_save

    def SaveExsitantFile(self):

        fname_Exfile = QFileDialog.getOpenFileName(self, 'Open File')
        address_Exfile = fname_Exfile[0]
        self.address_Exfile = address_Exfile
        wb_ex = openpyxl.load_workbook('%s' % address_Exfile)
        ws_ex = wb_ex.active

        textsetup_name = self.setup_name.text()
        textsetup_company = self.setup_company.text()
        textsetup_material = self.setup_company.text()
        textsetup_diameter = float(self.setup_diameter.text())
        textsetup_thickness = float(self.setup_thickness.text())
        textsetup_backth = float(self.setup_backth.text())

        textinput_sclength1 = float(self.input_sclength1.text())
        textinput_sclength2 = float(self.input_sclength2.text())
        textinput_fobs = float(self.input_fobs.text())
        textinput_wd = float(self.input_wd.text())
        textinput_bw = float(self.input_bw.text())
        textinput_fw = float(self.input_fw.text())
        textinput_velocity = float(self.input_velocity.text())

        if self.group_cosmetic.isChecked()==False:
            self.cosmetic_sctype = str(" ")
            textcosmetic_sclength1 = self.cosmetic_sclength1.setText(" ")
            textcosmetic_sclength2 = self.cosmetic_sclength2.setText(" ")

            textcosmetic_fobs = self.cosmetic_fobs.setText(" ")
            textcosmetic_wd = self.cosmetic_wd.setText(" ")
            textcosmetic_bw = self.cosmetic_bw.setText(" ")
            textcosmetic_fw = self.cosmetic_fw.setText(" ")
            textcosmetic_velocity = self.cosmetic_velocity.setText(" ")

        elif self.group_cosmetic.isChecked()==True:
            textcosmetic_sclength1 = float(self.cosmetic_sclength1.text())
            textcosmetic_sclength2 = float(self.cosmetic_sclength2.text())
            textcosmetic_fobs = float(self.cosmetic_fobs.text())
            textcosmetic_wd = float(self.cosmetic_wd.text())
            textcosmetic_bw = float(self.cosmetic_bw.text())
            textcosmetic_fw = float(self.cosmetic_fw.text())
            textcosmetic_velocity = float(self.cosmetic_velocity.text())

        textoutput_beadW = float(self.output_beadW.text())
        textoutput_beadH = float(self.output_beadH.text())
        textoutput_backW = float(self.output_backW.text())
        textoutput_backH = float(self.output_backH.text())
        textoutput_50W = float(self.output_50W.text())
        textoutput_50H = float(self.output_50H.text())
        textoutput_crack = float(self.output_crack.text())
        textoutput_porosity = float(self.output_porosity.text())

        allvariablelist = [textsetup_name, textsetup_company, self.setup_use, textsetup_material, self.setup_weldtype,
                           textsetup_diameter, textsetup_thickness, textsetup_backth, textinput_fobs,
                           textinput_wd, textinput_bw, textinput_fw, textinput_velocity, self.input_sctype,
                           textinput_sclength1, textinput_sclength2, textcosmetic_fobs, textcosmetic_wd,
                           textcosmetic_wd, textcosmetic_bw, textcosmetic_fw, textcosmetic_velocity,
                           self.cosmetic_sctype, textcosmetic_sclength1, textcosmetic_sclength2, textoutput_beadW,
                           textoutput_beadH, textoutput_backW, textoutput_backH, textoutput_50W, textoutput_50H,
                           textoutput_crack, textoutput_porosity]

        len(allvariablelist)
        row_ex = len(ws_ex['A']) + 1
        print("ok")
        for i in range(0, len(allvariablelist)):
            print("ok"+str(i))
            ws_ex.cell(row=row_ex, column=i + 1).value = allvariablelist[i]

        ws_ex.cell(row=row_ex, column=1).hyperlink = self.fname_openimage
        wb_ex.save('%s' % self.address_Exfile)

    def SaveNewFile(self):
        fname_Newfile = QFileDialog.getSaveFileName(self, 'Save File', '', ".xlsx(*.xlsx)")
        address_Newfile = fname_Newfile[0]
        self.address_Newfile = address_Newfile
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        print("ok1")

        textsetup_name = self.setup_name.text()
        textsetup_company = self.setup_company.text()
        textsetup_material = self.setup_company.text()
        textsetup_diameter = float(self.setup_diameter.text())
        textsetup_thickness = float(self.setup_thickness.text())
        textsetup_backth = float(self.setup_backth.text())

        textinput_sclength1 = float(self.input_sclength1.text())
        textinput_sclength2 = float(self.input_sclength2.text())
        textinput_fobs = float(self.input_fobs.text())
        textinput_wd = float(self.input_wd.text())
        textinput_bw = float(self.input_bw.text())
        textinput_fw = float(self.input_fw.text())
        textinput_velocity = float(self.input_velocity.text())

        if self.group_cosmetic.isChecked()==False:
            self.cosmetic_sctype = str(" ")
            textcosmetic_sclength1 = self.cosmetic_sclength1.setText(" ")
            textcosmetic_sclength2 = self.cosmetic_sclength2.setText(" ")

            textcosmetic_fobs = self.cosmetic_fobs.setText(" ")
            textcosmetic_wd = self.cosmetic_wd.setText(" ")
            textcosmetic_bw = self.cosmetic_bw.setText(" ")
            textcosmetic_fw = self.cosmetic_fw.setText(" ")
            textcosmetic_velocity = self.cosmetic_velocity.setText(" ")

        elif self.group_cosmetic.isChecked()==True:
            textcosmetic_sclength1 = float(self.cosmetic_sclength1.text())
            textcosmetic_sclength2 = float(self.cosmetic_sclength2.text())
            textcosmetic_fobs = float(self.cosmetic_fobs.text())
            textcosmetic_wd = float(self.cosmetic_wd.text())
            textcosmetic_bw = float(self.cosmetic_bw.text())
            textcosmetic_fw = float(self.cosmetic_fw.text())
            textcosmetic_velocity = float(self.cosmetic_velocity.text())

        textoutput_beadW = float(self.output_beadW.text())
        textoutput_beadH = float(self.output_beadH.text())
        textoutput_backW = float(self.output_backW.text())
        textoutput_backH = float(self.output_backH.text())
        textoutput_50W = float(self.output_50W.text())
        textoutput_50H = float(self.output_50H.text())
        textoutput_crack = float(self.output_crack.text())
        textoutput_porosity = float(self.output_porosity.text())

        allnamelist = ['Name', 'Company', 'Use', 'Material', 'Weld Type', 'Diameter', 'Thickness', 'Backing Thickness',
                       'Fobs', 'Wd', 'Bw', 'Fw', 'Velocity', 'S.C type',
                       'S.C length1', 'S.C length2', 'Cosmetic Fobs', 'Cosmetic Wd', 'Cosmetic Bw', 'Cosmetic Fw',
                       'Cosmetic velocity', 'Cosmetic S.C', 'Cosmetic S.C l1',
                       'Cosmetic S,C l2', 'Bead Width', 'Bead Height', 'Back Width', 'Back Height', '50%Width',
                       '50%Height', 'Crack', 'Porosity']
        allvariablelist = [textsetup_name, textsetup_company, self.setup_use, textsetup_material, self.setup_weldtype,
                           textsetup_diameter, textsetup_thickness, textsetup_backth, textinput_fobs,
                           textinput_wd, textinput_bw, textinput_fw, textinput_velocity, self.input_sctype,
                           textinput_sclength1, textinput_sclength2, textcosmetic_fobs, textcosmetic_wd,
                           textcosmetic_wd, textcosmetic_bw, textcosmetic_fw, textcosmetic_velocity,
                           self.cosmetic_sctype, textcosmetic_sclength1, textcosmetic_sclength2, textoutput_beadW,
                           textoutput_beadH, textoutput_backW, textoutput_backH, textoutput_50W, textoutput_50H,
                           textoutput_crack, textoutput_porosity]
        print("ok2")
        for i in range(0, len(allvariablelist)):
            ws_new.cell(row=1, column=i).value = allnamelist[i]
            ws_new.cell(row=2, column=i + 1).value = allvariablelist[i]
            print("ok3")

        ws_new.cell(row=2, column=1).hyperlink = self.fname_openimage
        wb_new.save('%s' % self.address_Newfile)

class TabCompare(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        name = QLabel('Name:')
        nameedit = QLineEdit()
        age = QLabel('Age:')
        ageedit = QLineEdit()
        nation = QLabel('Nation:')
        nationedit = QLineEdit()

        vbox = QVBoxLayout()
        vbox.addWidget(name)
        vbox.addWidget(nameedit)
        vbox.addWidget(age)
        vbox.addWidget(ageedit)
        vbox.addWidget(nation)
        vbox.addWidget(nationedit)
        vbox.addStretch()

        self.setLayout(vbox)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = MainWindow()
    sys.exit(app.exec_())
