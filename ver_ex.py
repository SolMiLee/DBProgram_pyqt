import sys
import xlwt
from PyQt5.QtWidgets import QMainWindow
from PyQt5.QtWidgets import QApplication
from PyQt5.QtWidgets import QVBoxLayout
from PyQt5.QtWidgets import QWidget
from PyQt5.QtWidgets import QLabel
from PyQt5.QtWidgets import QLineEdit
from PyQt5.QtWidgets import QPushButton
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtGui import QPixmap

import openpyxl
from openpyxl import Workbook


class window(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setGeometry(100, 100, 200, 200)
        winlay = QVBoxLayout()

        lb1 = QLabel("값1")
        self.lbe1 = QLineEdit(self)
        lbe_write = self.lbe1

        btn1 = QPushButton("이미지")
        btn1.clicked.connect(self.ImagebuttonClicked)
        self.lbl_im = QLabel()

        btn_save = QPushButton("save")
        btn_save.clicked.connect(self.saveButtonClicked)

        winlay.addWidget(lb1)
        winlay.addWidget(self.lbe1)
        winlay.addWidget(btn1)
        winlay.addWidget(self.lbl_im)
        winlay.addWidget(btn_save)

        winlay.addStretch(4)

        centralWidget = QWidget()
        centralWidget.setLayout(winlay)
        self.setCentralWidget(centralWidget)
        self.show()

    def ImagebuttonClicked(self):
        self.fname = QFileDialog.getOpenFileName(self, 'Open File')
        self.fnim = self.fname[0]
        pixmap = QPixmap(self.fname[0])
        self.lbl_im.setPixmap(pixmap)


        # print(self.fnim)
        #C:/Users/IWA-LSM/Desktop/EBW/DBProgram_pyqt/input.PNG

    def saveButtonClicked(self):
        fname_save = QFileDialog.getSaveFileName(self, 'Choose save path', '', '.xlsx(*.xlsx)')
        self.fnameee = fname_save[0]
        wb = openpyxl.Workbook()
        ws = wb.active
        lb_text = self.lbe1.text()
        ws.cell(row=1,column=1).value = lb_text
        ws.cell(row=1, column=1).hyperlink = self.fnim
        wb.save('%s'%self.fnameee)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = window()
    win.show()
    sys.exit(app.exec_())